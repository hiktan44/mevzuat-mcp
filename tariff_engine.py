"""Versioned, evidence-first tariff and trade-measure engine.

The engine only imports machine-readable archives discovered on official Ministry
of Trade landing pages.  Every parsed rate keeps its archive checksum, workbook,
sheet, row and source URL.  A rate is never guessed from prose or from an LLM.
"""

from __future__ import annotations

import asyncio
import hashlib
import html
import io
import json
import os
import re
import sqlite3
import time
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Literal
from urllib.parse import urljoin, urlsplit

import httpx
import xlrd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pydantic import BaseModel, Field, field_validator
from security_firewall import validate_outbound_url

_GTIP_RE = re.compile(r"^\d{4}(?:\d{2}){0,4}$")
_RATE_RE = re.compile(r"^-?\d+(?:[.,]\d+)?$")
_OFFICIAL_HOSTS = {"ticaret.gov.tr", "www.ticaret.gov.tr"}


def _now() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def _key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").casefold().replace("ı", "i"))
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _digits(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\D", "", str(value))


def _normalise_gtip(value: Any) -> str | None:
    digits = _digits(value)
    if len(digits) in {11}:
        digits = digits.zfill(12)
    if len(digits) in {4, 6, 8, 10, 12}:
        return digits
    return None


def _number(value: Any) -> tuple[float | None, str]:
    text = "" if value is None else str(value).strip().replace("%", "")
    if not text:
        return None, ""
    canonical = text.replace(" ", "").replace(",", ".")
    if not _RATE_RE.fullmatch(canonical):
        return None, text[:160]
    return float(canonical), text[:160]


def _official_url(url: str) -> bool:
    return (urlsplit(url).hostname or "").lower().rstrip(".") in _OFFICIAL_HOSTS


def _repair_zip_name(name: str) -> str:
    try:
        return name.encode("cp437").decode("cp857")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return name


def _zip_name(item: zipfile.ZipInfo) -> str:
    """Decode Ministry ZIP member names written with the Turkish DOS codepage."""
    return item.filename if item.flag_bits & 0x800 else _repair_zip_name(item.filename)


class TariffMeasure(BaseModel):
    gtip: str
    measure_type: Literal[
        "customs_duty",
        "additional_duty",
        "additional_financial_liability",
        "customs_duty_suspension",
        "customs_duty_end_use",
    ]
    rate: float | None = None
    rate_text: str
    country_group: str
    country_group_description: str
    footnote: str | None = None
    description: str | None = None
    condition: str | None = None
    list_name: str
    source_id: str
    source_title: str
    source_url: str
    archive_url: str
    archive_sha256: str
    source_file: str
    source_sheet: str
    source_row: int
    valid_from: str
    retrieved_at: str
    snapshot_id: str
    automatic_calculation_allowed: bool = True


class TariffSnapshot(BaseModel):
    id: str
    source_id: str
    source_title: str
    landing_url: str
    archive_url: str
    archive_sha256: str
    retrieved_at: str
    checked_at: str
    valid_from: str
    measure_count: int
    active: bool


class TariffLookupResult(BaseModel):
    status: Literal["matched", "partial", "not_found", "unavailable"]
    gtip: str
    match_mode: Literal["exact", "prefix"] = "exact"
    matched_gtips: list[str] = Field(default_factory=list)
    matched_gtip_count: int = 0
    origin_country: str | None = None
    resolved_country_group: str | None = None
    rate_variants: dict[str, list[float]] = Field(default_factory=dict)
    unambiguous_rates: dict[str, float] = Field(default_factory=dict)
    ambiguous_measure_types: list[str] = Field(default_factory=list)
    measures: list[TariffMeasure] = Field(default_factory=list)
    conditional_measures: list[TariffMeasure] = Field(default_factory=list)
    alternatives: list[TariffMeasure] = Field(default_factory=list)
    snapshots: list[TariffSnapshot] = Field(default_factory=list)
    measure_coverage: dict[str, "MeasureCoverage"] = Field(default_factory=dict)
    unresolved_measure_types: list[str] = Field(default_factory=list)
    warnings: list[str] = Field(default_factory=list)
    as_of: str


class MeasureCoverage(BaseModel):
    status: Literal["verified_snapshot", "partial_snapshot", "not_integrated", "user_confirmation_required"]
    source_ids: list[str] = Field(default_factory=list)
    note: str


class TariffTreeNode(BaseModel):
    """One deterministic branch in the HS6 -> CN8 -> TR10 -> GTIP12 tree."""

    code: str
    level: Literal["HS6", "CN8", "TR10", "GTIP12"]
    final: bool
    descendant_count: int = Field(..., ge=1)
    rate_status: Literal["unambiguous", "ambiguous", "origin_required"]
    unambiguous_rates: dict[str, float] = Field(default_factory=dict)
    rate_variants: dict[str, list[float]] = Field(default_factory=dict)
    ambiguous_measure_types: list[str] = Field(default_factory=list)
    warnings: list[str] = Field(default_factory=list)


class TariffDecisionTreeResult(BaseModel):
    """Deterministic descendants of a tariff prefix from active official tables."""

    status: Literal["matched", "not_found", "unavailable"]
    prefix: str
    level: Literal["HS6", "CN8", "TR10", "GTIP12"]
    next_level: Literal["CN8", "TR10", "GTIP12"] | None = None
    origin_country: str | None = None
    total_children: int = 0
    children: list[TariffTreeNode] = Field(default_factory=list)
    requires_user_selection: bool = True
    exact_gtip_selected: bool = False
    warnings: list[str] = Field(default_factory=list)
    as_of: str


class TariffSyncStatus(BaseModel):
    ready: bool
    syncing: bool
    last_checked_at: str | None = None
    active_snapshots: list[TariffSnapshot] = Field(default_factory=list)
    measure_count: int = 0
    errors: list[str] = Field(default_factory=list)
    sync_interval_seconds: int


class LandedCostInput(BaseModel):
    invoice_value: float = Field(..., gt=0, le=1_000_000_000)
    freight: float = Field(0, ge=0, le=1_000_000_000)
    insurance: float = Field(0, ge=0, le=1_000_000_000)
    other_costs: float = Field(0, ge=0, le=1_000_000_000)
    quantity: float | None = Field(None, gt=0, le=1_000_000_000)
    currency: str = Field("USD", min_length=3, max_length=3)
    customs_duty_rate: float | None = Field(None, ge=0, le=1000)
    additional_duty_rate: float | None = Field(None, ge=0, le=1000)
    additional_financial_liability_rate: float | None = Field(None, ge=0, le=1000)
    anti_dumping_amount: float | None = Field(None, ge=0, le=1_000_000_000)
    kkdf_rate: float | None = Field(None, ge=0, le=100)
    vat_rate: float | None = Field(None, ge=0, le=100)
    sct_amount: float | None = Field(None, ge=0, le=1_000_000_000)
    surveillance_unit_value: float | None = Field(None, ge=0, le=1_000_000_000)
    has_surveillance_certificate: bool | None = None

    @field_validator("currency")
    @classmethod
    def _currency(cls, value: str) -> str:
        value = value.upper()
        if not value.isalpha():
            raise ValueError("Para birimi üç harfli olmalıdır.")
        return value


class LandedCostResult(BaseModel):
    status: Literal["complete", "partial", "blocked"]
    currency: str
    lines: list[dict[str, Any]]
    customs_value: float
    vat_base: float | None = None
    landed_total: float | None = None
    unit_landed_cost: float | None = None
    missing_rates: list[str] = Field(default_factory=list)
    warnings: list[str] = Field(default_factory=list)
    formula_version: str = "tr-landed-cost-v2"


@dataclass(slots=True)
class _ParsedArchive:
    measures: list[dict[str, Any]]
    metadata: dict[str, Any]


_GROUP_DESCRIPTIONS = {
    "1": "AB, EFTA ve İthalat Rejimi Kararında 1 numaralı sütunda sayılan STA ülkeleri",
    "2": "Katar",
    "3": "Birleşik Arap Emirlikleri",
    "4": "En Az Gelişmiş Ülkeler",
    "5": "Özel Teşvik Düzenlemesinden Yararlanan Ülkeler",
    "6": "Gelişme Yolundaki Ülkeler",
    "7": "Diğer Ülkeler",
    "EAGÜ": "En Az Gelişmiş Ülkeler",
    "ÖTDÜ": "Özel Teşvik Düzenlemesinden Yararlanan Ülkeler",
    "GYÜ": "Gelişme Yolundaki Ülkeler",
    "DÜ": "Diğer Ülkeler",
}

_COLUMN_1_COUNTRIES = {
    "arnavutluk", "bolivarci venezuela cumhuriyeti", "venezuela", "birlesik krallik", "ingiltere",
    "bosna-hersek", "bosna hersek", "fas", "faroe adalari", "filistin", "gurcistan", "guney kore",
    "kore cumhuriyeti", "israil", "karadag", "kosova", "kuzey makedonya", "malezya", "misir",
    "morityus", "moldova", "sirbistan", "singapur", "sili", "tunus",
}
_EU_COUNTRIES = {
    "almanya", "avusturya", "belcika", "bulgaristan", "cekya", "cek cumhuriyeti", "danimarka",
    "estonya", "finlandiya", "fransa", "hirvatistan", "hollanda", "irlanda", "ispanya", "isvec",
    "italya", "kibris", "letonya", "litvanya", "luksemburg", "macaristan", "malta", "polonya",
    "portekiz", "romanya", "slovakya", "slovenya", "yunanistan",
}
_EFTA_COUNTRIES = {"izlanda", "lihtenstayn", "norvec", "isvicre"}
_EXPLICIT_LABELS = {
    "guney kore": "G.KORE", "kore cumhuriyeti": "G.KORE", "malezya": "MLZ",
    "singapur": "SNG", "kosova": "KOS", "iran": "İRAN", "venezuela": "VNZ",
    "birlesik arap emirlikleri": "BAE", "bae": "BAE", "gurcistan": "GÜR",
}


class TariffEngine:
    """Synchronise, query and diff official tariff snapshots."""

    def __init__(self, config_path: str | Path | None = None, data_dir: str | Path | None = None) -> None:
        config_file = Path(config_path or Path(__file__).with_name("tariff_sources.json"))
        config = json.loads(config_file.read_text(encoding="utf-8"))
        self.sources = list(config.get("sources", []))
        self.sync_interval_seconds = max(300, int(config.get("sync_interval_seconds", 21600)))
        default_root = Path(os.environ.get("XDG_CACHE_HOME", Path.home() / ".cache")) / "mevzuat-mcp"
        root = Path(data_dir or os.environ.get("MEVZUAT_DATA_DIR", default_root))
        root.mkdir(parents=True, exist_ok=True)
        self.db_path = root / "tariff.sqlite3"
        self._http = httpx.AsyncClient(
            follow_redirects=False,
            timeout=httpx.Timeout(60),
            headers={
                "User-Agent": "Gumrukce/1.4 (+official-tariff-sync)",
                "Accept-Language": "tr-TR,tr;q=0.9",
            },
            limits=httpx.Limits(max_connections=3, max_keepalive_connections=2),
        )
        self._sync_lock = asyncio.Lock()
        self._syncing = False
        self._errors: list[str] = []
        self._init_db()

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.db_path, timeout=30)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA journal_mode=WAL")
        connection.execute("PRAGMA foreign_keys=ON")
        return connection

    def _init_db(self) -> None:
        with self._connect() as db:
            db.executescript(
                """
                CREATE TABLE IF NOT EXISTS tariff_snapshots (
                    id TEXT PRIMARY KEY, source_id TEXT NOT NULL, source_title TEXT NOT NULL,
                    landing_url TEXT NOT NULL, archive_url TEXT NOT NULL, archive_sha256 TEXT NOT NULL,
                    retrieved_at TEXT NOT NULL, checked_at TEXT NOT NULL, valid_from TEXT NOT NULL,
                    measure_count INTEGER NOT NULL, active INTEGER NOT NULL DEFAULT 1,
                    metadata_json TEXT NOT NULL DEFAULT '{}'
                );
                CREATE INDEX IF NOT EXISTS idx_tariff_snapshots_source ON tariff_snapshots(source_id, active);
                CREATE TABLE IF NOT EXISTS tariff_measures (
                    id TEXT PRIMARY KEY, snapshot_id TEXT NOT NULL REFERENCES tariff_snapshots(id) ON DELETE CASCADE,
                    gtip TEXT NOT NULL, measure_type TEXT NOT NULL, rate REAL, rate_text TEXT NOT NULL,
                    country_group TEXT NOT NULL, country_group_description TEXT NOT NULL,
                    footnote TEXT, description TEXT, condition_text TEXT, list_name TEXT NOT NULL,
                    source_file TEXT NOT NULL, source_sheet TEXT NOT NULL, source_row INTEGER NOT NULL,
                    automatic_calculation_allowed INTEGER NOT NULL DEFAULT 1
                );
                CREATE INDEX IF NOT EXISTS idx_tariff_measures_gtip ON tariff_measures(gtip);
                CREATE INDEX IF NOT EXISTS idx_tariff_measures_snapshot ON tariff_measures(snapshot_id);
                """
            )

    async def close(self) -> None:
        await self._http.aclose()

    async def _get(self, url: str) -> httpx.Response:
        last: Exception | None = None
        for attempt in range(3):
            try:
                current_url = url
                for _ in range(6):
                    validate_outbound_url(current_url, allowed_hosts=_OFFICIAL_HOSTS)
                    response = await self._http.get(current_url)
                    if not response.is_redirect:
                        break
                    location = response.headers.get("location", "")
                    if not location:
                        raise ValueError("Resmî tarife yönlendirmesi hedefsiz")
                    current_url = urljoin(str(response.url), location)
                else:
                    raise ValueError("Resmî tarife kaynağı çok fazla yönlendirme yaptı")
                response.raise_for_status()
                return response
            except (httpx.HTTPError, httpx.TimeoutException, ValueError) as exc:
                last = exc
                await asyncio.sleep(0.5 * (attempt + 1))
        raise RuntimeError(f"Resmî kaynak alınamadı: {type(last).__name__}") from last

    async def _discover_archive(self, source: dict[str, Any]) -> str:
        response = await self._get(source["landing_url"])
        soup = BeautifulSoup(response.text, "lxml")
        wanted = _key(source.get("archive_text", ""))
        candidates: list[tuple[int, str]] = []
        for anchor in soup.select("a[href]"):
            href = urljoin(str(response.url), anchor.get("href", ""))
            if not _official_url(href) or ".zip" not in _key(urlsplit(href).path):
                continue
            label = _key(anchor.get_text(" ", strip=True))
            score = 100 if wanted and wanted in label else 10
            if str(datetime.now().year) in label or str(datetime.now().year) in href:
                score += 5
            candidates.append((score, href))
        if not candidates:
            raise ValueError("Resmî sayfada ZIP tablo bağlantısı bulunamadı.")
        return max(candidates, key=lambda item: item[0])[1]

    async def _download_archive(self, url: str) -> bytes:
        last: Exception | None = None
        for attempt in range(4):
            chunks: list[bytes] = []
            try:
                async with self._http.stream("GET", url, headers={"Referer": "https://ticaret.gov.tr/"}) as response:
                    response.raise_for_status()
                    async for chunk in response.aiter_bytes():
                        chunks.append(chunk)
                data = b"".join(chunks)
                with zipfile.ZipFile(io.BytesIO(data)) as archive:
                    if archive.testzip() is not None:
                        raise ValueError("Arşiv bütünlük denetiminden geçemedi.")
                return data
            except (httpx.HTTPError, httpx.TimeoutException, zipfile.BadZipFile, ValueError) as exc:
                # Some Ministry file endpoints reset the connection after sending
                # the complete body.  Accept it only when ZIP CRC validation passes.
                data = b"".join(chunks)
                if data:
                    try:
                        with zipfile.ZipFile(io.BytesIO(data)) as archive:
                            if archive.testzip() is None:
                                return data
                    except zipfile.BadZipFile:
                        pass
                last = exc
                await asyncio.sleep(0.75 * (attempt + 1))
        raise RuntimeError(f"Resmî tarife arşivi indirilemedi: {type(last).__name__}") from last

    @staticmethod
    def _sheet_rows(data: bytes, extension: str) -> Iterable[tuple[str, list[tuple[Any, ...]]]]:
        if extension == ".xlsx":
            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            for sheet in workbook.worksheets:
                yield sheet.title, [tuple(row) for row in sheet.iter_rows(values_only=True)]
        elif extension == ".xls":
            workbook = xlrd.open_workbook(file_contents=data)
            for sheet in workbook.sheets():
                yield sheet.name, [tuple(sheet.row_values(index)) for index in range(sheet.nrows)]

    @staticmethod
    def _docx_text(data: bytes) -> str:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            xml = archive.read("word/document.xml").decode("utf-8", errors="replace")
        xml = re.sub(r"</w:p[^>]*>", "\n", xml)
        xml = re.sub(r"<w:tab[^>]*/>", "\t", xml)
        return html.unescape(re.sub(r"<[^>]+>", "", xml))

    @staticmethod
    def _group_map(file_name: str, sheet_name: str, family: str) -> tuple[int, dict[int, str], str, str | None]:
        file_key, sheet_key = _key(file_name), _key(sheet_name)
        if family == "additional_duty":
            if "calisma" in sheet_key:
                return -1, {}, "", None
            if "ek-1" in file_key:
                return 0, {2: "1", 3: "2", 4: "3", 5: "4", 6: "5", 7: "6", 8: "7"}, "İGV Ek-1", None
            if "ek 2" in sheet_key:
                groups = {2: "AB/BK/B-HER/EFTA/F.ADA/G.KORE/MLZ", 3: "KOS", 4: "SNG", 5: "VNZ", 6: "BAE", 7: "EAGÜ", 8: "ÖTDÜ", 9: "GYÜ", 10: "DÜ"}
                return 0, groups, "İGV Ek-2", None
            if "ek 3" in sheet_key:
                groups = {2: "AB/BK/B-HER/EFTA/F.ADA", 3: "G.KORE", 4: "MLZ", 5: "SNG", 6: "KOS", 7: "İRAN", 8: "VNZ", 9: "BAE", 10: "EAGÜ", 11: "ÖTDÜ", 12: "GYÜ", 13: "DÜ"}
                return 0, groups, "İGV Ek-3", None
            return -1, {}, "", None

        if file_key.startswith(("ek-", "ek ")) or "tablo" in file_key:
            return -1, {}, "", None
        if "v say" in file_key:
            return 0, {5: "NİHAİ KULLANIM"}, "V Sayılı Liste", "customs_duty_suspension"
        if "vi say" in file_key:
            return 2, {5: "NİHAİ KULLANIM"}, "VI Sayılı Liste", "customs_duty_end_use"
        if "vii say" in file_key:
            return 0, {3: "NİHAİ KULLANIM"}, "VII Sayılı Liste", "customs_duty_end_use"
        if "iv say" in file_key:
            groups = {2: "EFTA/B-HER/F.ADA", 3: "AB/BK", 4: "G.KORE", 5: "MLZ", 6: "SNG", 7: "KOS", 8: "VNZ", 9: "BAE", 10: "TPS-OIC", 11: "D-8", 12: "DÜ", 13: "EFTA/F.ADA"}
            return 0, groups, "IV Sayılı Liste", None
        if "iii say" in file_key:
            groups = {2: "AB/BK/B-HER/EFTA/F.ADA", 3: "G.KORE", 4: "MLZ", 5: "SNG", 6: "KOS", 7: "İRAN", 8: "VNZ", 9: "BAE", 10: "EAGÜ", 11: "ÖTDÜ", 12: "GYÜ", 13: "DÜ"}
            return 0, groups, "III Sayılı Liste", None
        if "ii say" in file_key and "04-24" in file_key:
            groups = {2: "AB/B-HER/BK/EFTA/F.ADA/G.KORE/MLZ", 3: "KOS", 4: "SNG", 5: "VNZ", 6: "BAE", 7: "EAGÜ", 8: "ÖTDÜ", 9: "GYÜ", 10: "DÜ"}
            return 0, groups, "II Sayılı Liste (Tarım)", None
        if "ii say" in file_key:
            return 0, {2: "1", 3: "2", 4: "3", 5: "4", 6: "5", 7: "6", 8: "7"}, "II Sayılı Liste (Sanayi)", None
        if re.search(r"(^|\W)i say", file_key):
            groups = {2: "AB/BK", 3: "GÜR", 4: "B-HER", 5: "G.KORE", 6: "MLZ", 7: "SNG", 8: "KOS", 9: "VNZ", 10: "BAE", 11: "TPS-OIC", 12: "D-8", 13: "DÜ"}
            return 0, groups, "I Sayılı Liste", None
        return -1, {}, "", None

    @staticmethod
    def _country_metadata(archive: zipfile.ZipFile) -> dict[str, Any]:
        groups: dict[str, dict[str, str]] = {}
        sectors: dict[str, str] = {}
        for item in archive.infolist():
            if item.is_dir():
                continue
            filename = _zip_name(item)
            file_key = _key(filename)
            if file_key == "ek-1.xlsx":
                rows = next(TariffEngine._sheet_rows(archive.read(item), ".xlsx"))[1]
                current = ""
                for row in rows:
                    heading = _key(row[0] if row else "")
                    if "gelisme yolundaki" in heading:
                        current = "GYÜ"
                    elif "ozel tesvik" in heading:
                        current = "ÖTDÜ"
                    elif "en az gelismis" in heading:
                        current = "EAGÜ"
                    if not current:
                        continue
                    for country_col, exclusion_col in ((0, 1), (2, 3)):
                        if len(row) <= country_col:
                            continue
                        country = str(row[country_col] or "").strip()
                        country_key = _key(country)
                        if not country or "ulkeler" in country_key or country_key.startswith(("a-", "b-", "c-", "*")):
                            continue
                        groups[country_key] = {
                            "name": country,
                            "group": current,
                            "exclusions": str(row[exclusion_col] or "").strip() if len(row) > exclusion_col else "",
                        }
            elif file_key in {"ek-2.xls", "ek-3.xls", "ek-4.xls"}:
                extension = Path(filename).suffix.lower()
                for _, rows in TariffEngine._sheet_rows(archive.read(item), extension):
                    for row in rows:
                        if len(row) < 3:
                            continue
                        gtip = _normalise_gtip(row[2])
                        sector = str(row[0] or "").strip()
                        if gtip and sector:
                            sectors[f"{file_key}:{gtip}"] = sector
        return {"gts_countries": groups, "gts_sectors": sectors}

    @classmethod
    def _parse_archive(cls, data: bytes, source: dict[str, Any], archive_url: str, sha256: str, retrieved_at: str) -> _ParsedArchive:
        measures: list[dict[str, Any]] = []
        metadata: dict[str, Any] = {}
        family = source["measure_family"]
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            if family == "customs_duty":
                metadata = cls._country_metadata(archive)
            for item in archive.infolist():
                if item.is_dir():
                    continue
                filename = _zip_name(item)
                extension = Path(filename).suffix.lower()
                if extension not in {".xlsx", ".xls"}:
                    continue
                for sheet_name, rows in cls._sheet_rows(archive.read(item), extension):
                    code_col, group_columns, list_name, override_type = cls._group_map(filename, sheet_name, family)
                    if code_col < 0:
                        continue
                    for row_number, row in enumerate(rows, 1):
                        if len(row) <= code_col:
                            continue
                        gtip = _normalise_gtip(row[code_col])
                        if not gtip:
                            continue
                        footnote = str(row[code_col + 1] or "").strip()[:500] if len(row) > code_col + 1 else ""
                        description = ""
                        if override_type == "customs_duty_end_use":
                            description_col = 4 if code_col == 2 else 1
                            if len(row) > description_col:
                                description = str(row[description_col] or "").strip()[:2000]
                        elif override_type == "customs_duty_suspension" and len(row) > 4:
                            description = str(row[4] or "").strip()[:2000]
                        for column, group in group_columns.items():
                            if len(row) <= column:
                                continue
                            rate, rate_text = _number(row[column])
                            if not rate_text:
                                continue
                            measure_type = override_type or family
                            if family == "customs_duty" and list_name == "IV Sayılı Liste" and column == 13:
                                measure_type = "additional_financial_liability"
                            automatic = rate is not None and override_type is None and not footnote
                            condition = None
                            if override_type:
                                condition = "Bu oran yalnızca ilgili nihai kullanım/askıya alma şartları sağlanırsa uygulanabilir."
                            elif footnote:
                                condition = "Dipnot koşulu resmî karar ekinden doğrulanmalıdır."
                            measures.append(
                                {
                                    "gtip": gtip,
                                    "measure_type": measure_type,
                                    "rate": rate,
                                    "rate_text": rate_text,
                                    "country_group": group,
                                    "country_group_description": _GROUP_DESCRIPTIONS.get(group, group.replace("/", ", ")),
                                    "footnote": footnote or None,
                                    "description": description or None,
                                    "condition": condition,
                                    "list_name": list_name,
                                    "source_file": filename,
                                    "source_sheet": sheet_name,
                                    "source_row": row_number,
                                    "automatic_calculation_allowed": automatic,
                                }
                            )
        return _ParsedArchive(measures=measures, metadata=metadata)

    def _latest_snapshot_row(self, source_id: str | None = None) -> sqlite3.Row | None:
        with self._connect() as db:
            if source_id:
                return db.execute(
                    "SELECT * FROM tariff_snapshots WHERE source_id=? ORDER BY checked_at DESC LIMIT 1",
                    (source_id,),
                ).fetchone()
            return db.execute("SELECT * FROM tariff_snapshots ORDER BY checked_at DESC LIMIT 1").fetchone()

    async def _sync_source(self, source: dict[str, Any]) -> TariffSnapshot:
        archive_url = await self._discover_archive(source)
        data = await self._download_archive(archive_url)
        checksum = hashlib.sha256(data).hexdigest()
        snapshot_id = f"{source['id']}:{checksum[:20]}"
        checked_at = _now()
        with self._connect() as db:
            existing = db.execute("SELECT * FROM tariff_snapshots WHERE id=?", (snapshot_id,)).fetchone()
            if existing:
                db.execute("UPDATE tariff_snapshots SET checked_at=?, active=1 WHERE id=?", (checked_at, snapshot_id))
                db.execute("UPDATE tariff_snapshots SET active=0 WHERE source_id=? AND id<>?", (source["id"], snapshot_id))
                db.commit()
                return self._snapshot(existing, checked_at=checked_at, active=True)

        retrieved_at = checked_at
        parsed = await asyncio.to_thread(self._parse_archive, data, source, archive_url, checksum, retrieved_at)
        with self._connect() as db:
            db.execute("UPDATE tariff_snapshots SET active=0 WHERE source_id=?", (source["id"],))
            db.execute(
                """INSERT INTO tariff_snapshots
                (id,source_id,source_title,landing_url,archive_url,archive_sha256,retrieved_at,checked_at,valid_from,measure_count,active,metadata_json)
                VALUES (?,?,?,?,?,?,?,?,?,?,1,?)""",
                (
                    snapshot_id, source["id"], source["title"], source["landing_url"], archive_url, checksum,
                    retrieved_at, checked_at, source["valid_from"], len(parsed.measures),
                    json.dumps(parsed.metadata, ensure_ascii=False, separators=(",", ":")),
                ),
            )
            for item in parsed.measures:
                identity = hashlib.sha256(
                    json.dumps([snapshot_id, item], sort_keys=True, ensure_ascii=False).encode("utf-8")
                ).hexdigest()
                db.execute(
                    """INSERT INTO tariff_measures
                    (id,snapshot_id,gtip,measure_type,rate,rate_text,country_group,country_group_description,
                     footnote,description,condition_text,list_name,source_file,source_sheet,source_row,automatic_calculation_allowed)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        identity, snapshot_id, item["gtip"], item["measure_type"], item["rate"], item["rate_text"],
                        item["country_group"], item["country_group_description"], item["footnote"], item["description"],
                        item["condition"], item["list_name"], item["source_file"], item["source_sheet"],
                        item["source_row"], int(item["automatic_calculation_allowed"]),
                    ),
                )
            db.commit()
            row = db.execute("SELECT * FROM tariff_snapshots WHERE id=?", (snapshot_id,)).fetchone()
        return self._snapshot(row)

    async def sync(self, *, force: bool = False) -> TariffSyncStatus:
        async with self._sync_lock:
            latest = self._latest_snapshot_row()
            if not force and latest:
                age = time.time() - datetime.fromisoformat(latest["checked_at"]).timestamp()
                if age < self.sync_interval_seconds and self.status().ready:
                    return self.status()
            self._syncing = True
            self._errors = []
            try:
                for source in self.sources:
                    try:
                        await self._sync_source(source)
                    except Exception as exc:
                        self._errors.append(f"{source['id']}: {type(exc).__name__}: {exc}")
            finally:
                self._syncing = False
            return self.status()

    async def periodic_sync_loop(self) -> None:
        """Refresh at startup and on the configured cadence without blocking ASGI startup."""
        while True:
            try:
                await self.sync()
            except asyncio.CancelledError:
                raise
            except Exception as exc:
                self._errors.append(f"periodic_sync: {type(exc).__name__}: {exc}")
            await asyncio.sleep(self.sync_interval_seconds)

    @staticmethod
    def _snapshot(row: sqlite3.Row, *, checked_at: str | None = None, active: bool | None = None) -> TariffSnapshot:
        return TariffSnapshot(
            id=row["id"], source_id=row["source_id"], source_title=row["source_title"], landing_url=row["landing_url"],
            archive_url=row["archive_url"], archive_sha256=row["archive_sha256"], retrieved_at=row["retrieved_at"],
            checked_at=checked_at or row["checked_at"], valid_from=row["valid_from"], measure_count=row["measure_count"],
            active=bool(row["active"]) if active is None else active,
        )

    def status(self) -> TariffSyncStatus:
        with self._connect() as db:
            rows = db.execute("SELECT * FROM tariff_snapshots WHERE active=1 ORDER BY source_id").fetchall()
            count = db.execute(
                "SELECT COUNT(*) FROM tariff_measures m JOIN tariff_snapshots s ON s.id=m.snapshot_id WHERE s.active=1"
            ).fetchone()[0]
        snapshots = [self._snapshot(row) for row in rows]
        return TariffSyncStatus(
            ready=bool(snapshots) and all(item.measure_count > 0 for item in snapshots),
            syncing=self._syncing,
            last_checked_at=max((item.checked_at for item in snapshots), default=None),
            active_snapshots=snapshots,
            measure_count=count,
            errors=list(self._errors),
            sync_interval_seconds=self.sync_interval_seconds,
        )

    @staticmethod
    def _matching_group(origin: str, labels: set[str], metadata: dict[str, Any], gtip: str) -> tuple[str | None, list[str]]:
        origin_key = _key(origin).strip()
        warnings: list[str] = []
        if not origin_key:
            return None, warnings
        if origin_key in _EU_COUNTRIES or origin_key in _EFTA_COUNTRIES or origin_key in _COLUMN_1_COUNTRIES:
            if "1" in labels:
                return "1", warnings
            for label in labels:
                if any(token in label for token in ("AB", "EFTA")):
                    return label, warnings
        if origin_key in {"katar", "qatar"} and "2" in labels:
            return "2", warnings
        if origin_key in {"birlesik arap emirlikleri", "bae", "united arab emirates"}:
            if "3" in labels:
                return "3", warnings
            if "BAE" in labels:
                return "BAE", warnings
        explicit = _EXPLICIT_LABELS.get(origin_key)
        if explicit:
            for label in labels:
                if explicit in label:
                    return label, warnings

        gts = metadata.get("gts_countries", {}).get(origin_key)
        if gts:
            group = gts["group"]
            exclusions = gts.get("exclusions", "")
            if exclusions:
                sector_map = metadata.get("gts_sectors", {})
                prefix = {"GYÜ": "ek-2.xls", "ÖTDÜ": "ek-3.xls", "EAGÜ": "ek-4.xls"}[group]
                sector = sector_map.get(f"{prefix}:{gtip}")
                if sector and any(_key(part).strip() == _key(sector).strip() for part in exclusions.split(",")):
                    warnings.append(f"{origin} için {sector} sektörü GTS istisnasında; diğer ülkeler sütunu kullanıldı.")
                    return "7" if "7" in labels else ("DÜ" if "DÜ" in labels else None), warnings
                warnings.append(f"{origin} GTS ülkesidir; ürün sektörü istisnası ‘{exclusions}’ ayrıca doğrulanmalıdır.")
            if group in labels:
                return group, warnings
            numeric = {"EAGÜ": "4", "ÖTDÜ": "5", "GYÜ": "6"}[group]
            if numeric in labels:
                return numeric, warnings

        # Only use the residual group when the origin is not a listed preference
        # country in the official snapshot. This is a deterministic set complement.
        if "7" in labels:
            return "7", warnings
        if "DÜ" in labels:
            return "DÜ", warnings
        return None, warnings

    def _measure(self, row: sqlite3.Row, snapshot: sqlite3.Row) -> TariffMeasure:
        return TariffMeasure(
            gtip=row["gtip"], measure_type=row["measure_type"], rate=row["rate"], rate_text=row["rate_text"],
            country_group=row["country_group"], country_group_description=row["country_group_description"],
            footnote=row["footnote"], description=row["description"], condition=row["condition_text"],
            list_name=row["list_name"], source_id=snapshot["source_id"], source_title=snapshot["source_title"],
            source_url=snapshot["landing_url"], archive_url=snapshot["archive_url"],
            archive_sha256=snapshot["archive_sha256"], source_file=_repair_zip_name(row["source_file"]), source_sheet=row["source_sheet"],
            source_row=row["source_row"], valid_from=snapshot["valid_from"], retrieved_at=snapshot["retrieved_at"],
            snapshot_id=snapshot["id"], automatic_calculation_allowed=bool(row["automatic_calculation_allowed"]),
        )

    @staticmethod
    def _measure_coverage(snapshots: list[sqlite3.Row]) -> dict[str, MeasureCoverage]:
        active = {str(snapshot["source_id"]) for snapshot in snapshots}
        return {
            "customs_duty": MeasureCoverage(
                status="verified_snapshot" if "import_regime" in active else "not_integrated",
                source_ids=["import_regime"] if "import_regime" in active else [],
                note="Menşe sütunu, dipnot ve alt GTİP birliği sağlanırsa otomatik kullanılabilir.",
            ),
            "additional_duty": MeasureCoverage(
                status="verified_snapshot" if "additional_duty" in active else "not_integrated",
                source_ids=["additional_duty"] if "additional_duty" in active else [],
                note="İGV, menşe ve bütün GTİP12 alt satırları aynı oranı verirse otomatik kullanılabilir.",
            ),
            "additional_financial_liability": MeasureCoverage(
                status="partial_snapshot" if "import_regime" in active else "not_integrated",
                source_ids=["import_regime"] if "import_regime" in active else [],
                note="Konsolide cetveldeki IV sayılı liste okunur; diğer ek mali yükümlülük kararları ayrıca doğrulanmalıdır.",
            ),
            "anti_dumping": MeasureCoverage(
                status="not_integrated",
                note="Ürün, menşe ve üretici/ihracatçı bazlı güncel damping/sübvansiyon önlemi henüz yapılandırılmış hesap motorunda değildir.",
            ),
            "surveillance": MeasureCoverage(
                status="not_integrated",
                note="Gözetim tebliği, birim kıymet, menşe ve yürürlük tarihi ayrıca doğrulanmalıdır.",
            ),
            "safeguard": MeasureCoverage(
                status="not_integrated",
                note="Korunma önlemi ve varsa ülke/istisna kapsamı ayrıca doğrulanmalıdır.",
            ),
            "tariff_quota": MeasureCoverage(
                status="not_integrated",
                note="Tarife kontenjanı tahsis ve bakiye durumu işlem tarihinde ayrıca doğrulanmalıdır.",
            ),
            "vat": MeasureCoverage(
                status="user_confirmation_required",
                note="Ürüne özgü güncel KDV oranı resmî kaynaktan doğrulanıp girilmelidir.",
            ),
            "kkdf": MeasureCoverage(
                status="user_confirmation_required",
                note="Ödeme şekli ve istisnaya göre KKDF oranı doğrulanıp girilmelidir.",
            ),
            "sct": MeasureCoverage(
                status="user_confirmation_required",
                note="ÖTV kapsamı ve matrahı doğrulanıp toplam tutar girilmelidir.",
            ),
        }

    async def lookup(self, gtip: str, *, origin_country: str | None = None, auto_sync: bool = True) -> TariffLookupResult:
        normalised = _normalise_gtip(gtip)
        if not normalised or len(normalised) not in {6, 8, 10, 12}:
            raise ValueError("Tarife sorgusu için 6, 8, 10 veya 12 haneli HS/CN/GTİP kodu gereklidir.")
        match_mode: Literal["exact", "prefix"] = "exact" if len(normalised) == 12 else "prefix"
        if auto_sync and not self.status().ready:
            await self.sync()
        with self._connect() as db:
            snapshots = db.execute("SELECT * FROM tariff_snapshots WHERE active=1 ORDER BY source_id").fetchall()
            if not snapshots:
                return TariffLookupResult(status="unavailable", gtip=normalised, origin_country=origin_country, as_of=_now(), warnings=["Resmî tarife tabloları henüz eşitlenmedi."])
            all_rows: list[tuple[sqlite3.Row, sqlite3.Row]] = []
            metadata: dict[str, Any] = {}
            for snapshot in snapshots:
                metadata.update(json.loads(snapshot["metadata_json"] or "{}"))
                if match_mode == "exact":
                    rows = db.execute(
                        "SELECT * FROM tariff_measures WHERE snapshot_id=? AND gtip=? ORDER BY gtip,measure_type,list_name,country_group",
                        (snapshot["id"], normalised),
                    ).fetchall()
                else:
                    rows = db.execute(
                        "SELECT * FROM tariff_measures WHERE snapshot_id=? AND gtip LIKE ? ORDER BY gtip,measure_type,list_name,country_group",
                        (snapshot["id"], f"{normalised}%"),
                    ).fetchall()
                all_rows.extend((row, snapshot) for row in rows)
        if not all_rows:
            coverage = self._measure_coverage(snapshots)
            return TariffLookupResult(
                status="not_found", gtip=normalised, match_mode=match_mode, origin_country=origin_country,
                snapshots=[self._snapshot(row) for row in snapshots], measure_coverage=coverage,
                unresolved_measure_types=[key for key, item in coverage.items() if item.status != "verified_snapshot"],
                as_of=_now(),
                warnings=["Bu kod aktif resmî tarife/İGV tablolarında bulunamadı; kod ve fasıl doğrulaması gerekir."],
            )

        matched_gtips = sorted({row["gtip"] for row, _ in all_rows})
        selected_by_scope: dict[tuple[str, str], str | None] = {}
        warnings: list[str] = []
        scopes = {(row["gtip"], snapshot["id"]) for row, snapshot in all_rows}
        for matched_gtip, snapshot_id in sorted(scopes):
            labels = {
                row["country_group"]
                for row, snapshot in all_rows
                if row["gtip"] == matched_gtip and snapshot["id"] == snapshot_id
            }
            selected_group, group_warnings = self._matching_group(
                origin_country or "", labels, metadata, matched_gtip
            )
            selected_by_scope[(matched_gtip, snapshot_id)] = selected_group
            warnings.extend(group_warnings)
        warnings = list(dict.fromkeys(warnings))
        selected_groups = {group for group in selected_by_scope.values() if group}
        selected = " / ".join(sorted(selected_groups)) if selected_groups else None
        if len(selected_groups) > 1:
            warnings.append(
                "Resmî tablolar aynı menşe grubu için farklı sütun etiketleri kullanıyor: "
                + " / ".join(sorted(selected_groups))
            )
        if any(int(str(snapshot["valid_from"])[:4]) < datetime.now().year for snapshot in snapshots):
            warnings.append(
                "Aktif tarife snapshot'ı cari yıldan eskidir; oran otomatik karar için kullanılmadan önce yıllık cetvel güncellemesi doğrulanmalıdır."
            )
        primary: list[TariffMeasure] = []
        conditional: list[TariffMeasure] = []
        alternatives: list[TariffMeasure] = []
        for row, snapshot in all_rows:
            measure = self._measure(row, snapshot)
            if measure.measure_type in {"customs_duty_suspension", "customs_duty_end_use"}:
                conditional.append(measure)
            elif selected_by_scope.get((measure.gtip, measure.snapshot_id)) == measure.country_group:
                primary.append(measure)
            else:
                alternatives.append(measure)
        if not origin_country:
            warnings.append("Menşe ülke verilmediği için uygulanacak ülke sütunu seçilmedi.")
        elif not selected:
            warnings.append("Menşe ülke resmî ülke gruplarına güvenle eşlenemedi; oran otomatik seçilmedi.")
        if any(item.footnote for item in primary):
            warnings.append("Seçilen oranlardan en az biri dipnotludur; otomatik maliyet hesabından önce dipnot şartı doğrulanmalıdır.")

        calculable_types = ("customs_duty", "additional_duty", "additional_financial_liability")
        rate_variants: dict[str, list[float]] = {}
        unambiguous_rates: dict[str, float] = {}
        ambiguous_measure_types: list[str] = []
        matched_gtip_set = set(matched_gtips)
        for measure_type in calculable_types:
            typed = [item for item in primary if item.measure_type == measure_type]
            if not typed:
                continue
            valid = [
                item for item in typed
                if item.automatic_calculation_allowed and item.rate is not None and not item.footnote
            ]
            values = sorted({float(item.rate) for item in valid if item.rate is not None})
            rate_variants[measure_type] = values
            covered_gtips = {item.gtip for item in valid}
            has_unsafe_row = len(valid) != len(typed)
            if len(values) == 1 and covered_gtips == matched_gtip_set and not has_unsafe_row:
                unambiguous_rates[measure_type] = values[0]
            else:
                ambiguous_measure_types.append(measure_type)

        if match_mode == "prefix":
            warnings.insert(
                0,
                f"{len(normalised)} haneli kod {len(matched_gtips)} adet 12 haneli Türk GTİP satırıyla eşleşti.",
            )
            if ambiguous_measure_types:
                warnings.append(
                    "Alt GTİP satırlarında oran veya önlemin varlığı değişiyor; kesin alt GTİP seçilmeden bu kalemler maliyete otomatik alınmadı."
                )
            elif unambiguous_rates:
                warnings.append(
                    "Gösterilen otomatik oranlar bütün eşleşen 12 haneli satırlarda aynıdır; beyan öncesinde kesin 12 haneli GTİP yine doğrulanmalıdır."
                )
        coverage = self._measure_coverage(snapshots)
        unresolved = [key for key, item in coverage.items() if item.status != "verified_snapshot"]
        if unresolved:
            warnings.append(
                "Bu sonuç bütün ticaret politikası ve iç vergi kalemlerinin doğrulandığı anlamına gelmez; "
                "kapsam matrisi 'partial/not_integrated/user_confirmation_required' kalemlerini ayrı gösterir."
            )
        return TariffLookupResult(
            status="matched" if primary and not ambiguous_measure_types else "partial",
            gtip=normalised, match_mode=match_mode, matched_gtips=matched_gtips[:500],
            matched_gtip_count=len(matched_gtips), origin_country=origin_country,
            resolved_country_group=selected, rate_variants=rate_variants,
            unambiguous_rates=unambiguous_rates, ambiguous_measure_types=ambiguous_measure_types,
            measures=primary[:500], conditional_measures=conditional[:240],
            alternatives=alternatives[:120], snapshots=[self._snapshot(row) for row in snapshots],
            measure_coverage=coverage, unresolved_measure_types=unresolved, warnings=warnings, as_of=_now(),
        )

    async def decision_tree(
        self,
        gtip: str,
        *,
        origin_country: str | None = None,
        auto_sync: bool = True,
    ) -> TariffDecisionTreeResult:
        """Return the next official tariff level without guessing a child code.

        The method deliberately exposes every immediate branch.  It never ranks or
        auto-selects a child, because a rate row proving that a code exists does not
        prove that the user's goods belong under that code.
        """
        normalised = _normalise_gtip(gtip)
        level_by_length: dict[int, Literal["HS6", "CN8", "TR10", "GTIP12"]] = {
            6: "HS6",
            8: "CN8",
            10: "TR10",
            12: "GTIP12",
        }
        next_by_length: dict[int, tuple[int, Literal["CN8", "TR10", "GTIP12"]]] = {
            6: (8, "CN8"),
            8: (10, "TR10"),
            10: (12, "GTIP12"),
        }
        if not normalised or len(normalised) not in level_by_length:
            raise ValueError("Tarife karar ağacı için 6, 8, 10 veya 12 haneli kod gereklidir.")
        if auto_sync and not self.status().ready:
            await self.sync()
        if not self.status().ready:
            return TariffDecisionTreeResult(
                status="unavailable",
                prefix=normalised,
                level=level_by_length[len(normalised)],
                origin_country=origin_country,
                warnings=["Resmî tarife tabloları henüz eşitlenmedi."],
                as_of=_now(),
            )

        current = await self.lookup(normalised, origin_country=origin_country, auto_sync=False)
        if current.status in {"not_found", "unavailable"} or current.matched_gtip_count < 1:
            return TariffDecisionTreeResult(
                status="not_found" if current.status == "not_found" else "unavailable",
                prefix=normalised,
                level=level_by_length[len(normalised)],
                origin_country=origin_country,
                warnings=current.warnings,
                as_of=current.as_of,
            )

        if len(normalised) == 12:
            return TariffDecisionTreeResult(
                status="matched",
                prefix=normalised,
                level="GTIP12",
                origin_country=origin_country,
                requires_user_selection=False,
                exact_gtip_selected=True,
                warnings=current.warnings,
                as_of=current.as_of,
            )

        child_length, next_level = next_by_length[len(normalised)]
        child_codes = sorted({code[:child_length] for code in current.matched_gtips})
        child_lookups = await asyncio.gather(
            *(
                self.lookup(code, origin_country=origin_country, auto_sync=False)
                for code in child_codes
            )
        )
        children: list[TariffTreeNode] = []
        for code, lookup in zip(child_codes, child_lookups):
            if not origin_country:
                rate_status: Literal["unambiguous", "ambiguous", "origin_required"] = "origin_required"
            elif lookup.ambiguous_measure_types or "customs_duty" not in lookup.unambiguous_rates:
                rate_status = "ambiguous"
            else:
                rate_status = "unambiguous"
            children.append(
                TariffTreeNode(
                    code=code,
                    level=next_level,
                    final=child_length == 12,
                    descendant_count=lookup.matched_gtip_count,
                    rate_status=rate_status,
                    unambiguous_rates=lookup.unambiguous_rates,
                    rate_variants=lookup.rate_variants,
                    ambiguous_measure_types=lookup.ambiguous_measure_types,
                    warnings=lookup.warnings[:6],
                )
            )

        warnings = [
            f"{normalised} altında {len(children)} adet {next_level} dalı ve "
            f"{current.matched_gtip_count} adet GTİP12 satırı bulundu.",
            "Bir alt dal yalnızca kullanıcı tarafından doğrulanan ürün evsafıyla seçilmelidir.",
        ]
        if len(children) == 1:
            warnings.append(
                "Tek alt dal bulunması ürün sınıflandırmasını hukuken kesinleştirmez; ürün evsafı yine doğrulanmalıdır."
            )
        return TariffDecisionTreeResult(
            status="matched",
            prefix=normalised,
            level=level_by_length[len(normalised)],
            next_level=next_level,
            origin_country=origin_country,
            total_children=len(children),
            children=children,
            requires_user_selection=True,
            exact_gtip_selected=False,
            warnings=warnings,
            as_of=current.as_of,
        )

    async def calculate(
        self,
        gtip: str,
        origin_country: str,
        data: LandedCostInput,
    ) -> dict[str, Any]:
        """Apply only one unambiguous, unfootnoted official rate per measure type."""
        lookup = await self.lookup(gtip, origin_country=origin_country)
        safe_rates = lookup.unambiguous_rates
        conflicts = [
            f"{measure_type}: alt GTİP satırlarında oran veya önlem kapsamı farklı"
            for measure_type in lookup.ambiguous_measure_types
        ]
        enriched = data.model_copy(
            update={
                "customs_duty_rate": data.customs_duty_rate if data.customs_duty_rate is not None else safe_rates.get("customs_duty"),
                "additional_duty_rate": data.additional_duty_rate if data.additional_duty_rate is not None else safe_rates.get("additional_duty"),
                "additional_financial_liability_rate": (
                    data.additional_financial_liability_rate
                    if data.additional_financial_liability_rate is not None
                    else safe_rates.get("additional_financial_liability")
                ),
            }
        )
        cost = calculate_landed_cost(enriched)
        cost.warnings.extend(conflicts)
        return {
            "tariff": lookup.model_dump(mode="json"),
            "cost": cost.model_dump(mode="json"),
            "legal_notice": (
                "Hesap, gösterilen snapshot ve kullanıcı girdileriyle hazırlanmış ön çalışmadır. "
                "GTİP, menşe, kıymet, belge ve yürürlük durumu gümrük işlemi öncesinde doğrulanmalıdır."
            ),
        }

    def changes(self, source_id: str, *, limit: int = 200) -> dict[str, Any]:
        with self._connect() as db:
            snapshots = db.execute(
                "SELECT * FROM tariff_snapshots WHERE source_id=? ORDER BY retrieved_at DESC LIMIT 2", (source_id,)
            ).fetchall()
            if len(snapshots) < 2:
                return {"source_id": source_id, "status": "no_previous_snapshot", "changes": []}
            newer, older = snapshots
            def rows(snapshot_id: str) -> dict[tuple[str, str, str], sqlite3.Row]:
                result = db.execute("SELECT * FROM tariff_measures WHERE snapshot_id=?", (snapshot_id,)).fetchall()
                return {(row["gtip"], row["measure_type"], row["country_group"]): row for row in result}
            current, previous = rows(newer["id"]), rows(older["id"])
        changes: list[dict[str, Any]] = []
        for key in sorted(set(current) | set(previous)):
            new, old = current.get(key), previous.get(key)
            before = old["rate_text"] if old else None
            after = new["rate_text"] if new else None
            before_note = old["footnote"] if old else None
            after_note = new["footnote"] if new else None
            if before != after or before_note != after_note:
                changes.append({"gtip": key[0], "measure_type": key[1], "country_group": key[2], "before": before, "after": after, "before_footnote": before_note, "after_footnote": after_note})
        return {
            "source_id": source_id, "status": "compared", "new_snapshot": newer["id"], "old_snapshot": older["id"],
            "total_changes": len(changes), "changes": changes[: max(1, min(limit, 1000))],
        }


def calculate_landed_cost(data: LandedCostInput) -> LandedCostResult:
    """Reproducible arithmetic; no tax rate is inferred inside this function."""
    warnings: list[str] = []
    missing: list[str] = []
    customs_value = data.invoice_value + data.freight + data.insurance
    if data.surveillance_unit_value is None:
        missing.append("Gözetim birim kıymeti (uygulanmıyorsa 0)")
    elif data.surveillance_unit_value > 0:
        if data.quantity is None:
            missing.append("Gözetim kıymeti için miktar")
        elif data.has_surveillance_certificate is False:
            customs_value = max(customs_value, data.surveillance_unit_value * data.quantity)
            warnings.append("Gözetim belgesi olmadığı beyanına göre gümrük kıymeti birim gözetim kıymetine yükseltildi.")
        elif data.has_surveillance_certificate is None:
            missing.append("Gözetim belgesi durumu")

    lines: list[dict[str, Any]] = [
        {"code": "customs_value", "label": "Gümrük kıymeti", "base": None, "rate": None, "amount": round(customs_value, 2), "formula": "fatura + navlun + sigorta (varsa gözetim düzeltmesi)"}
    ]

    def percentage(code: str, label: str, base: float, rate: float | None) -> float | None:
        if rate is None:
            missing.append(label)
            lines.append({"code": code, "label": label, "base": round(base, 2), "rate": None, "amount": None, "formula": "oran doğrulanmadı"})
            return None
        amount = base * rate / 100
        lines.append({"code": code, "label": label, "base": round(base, 2), "rate": rate, "amount": round(amount, 2), "formula": f"{round(base, 2)} × %{rate}"})
        return amount

    duty = percentage("customs_duty", "Gümrük vergisi oranı", customs_value, data.customs_duty_rate)
    additional = percentage("additional_duty", "İlave gümrük vergisi oranı", customs_value, data.additional_duty_rate)
    emy = percentage("financial_liability", "Ek mali yükümlülük oranı", customs_value, data.additional_financial_liability_rate)
    if data.anti_dumping_amount is None:
        missing.append("Damping/sübvansiyon önlemi (uygulanmıyorsa 0)")
        lines.append({"code": "anti_dumping", "label": "Damping karşıtı vergi", "base": None, "rate": None, "amount": None, "formula": "uygulanabilirlik doğrulanmadı"})
    else:
        lines.append({"code": "anti_dumping", "label": "Damping karşıtı vergi", "base": None, "rate": None, "amount": round(data.anti_dumping_amount, 2), "formula": "doğrulanmış sabit/toplam tutar"})
    kkdf = percentage("kkdf", "KKDF oranı", data.invoice_value, data.kkdf_rate)
    if data.sct_amount is None:
        missing.append("ÖTV tutarı (uygulanmıyorsa 0)")
        lines.append({"code": "sct", "label": "ÖTV", "base": None, "rate": None, "amount": None, "formula": "uygulanabilirlik doğrulanmadı"})
    else:
        lines.append({"code": "sct", "label": "ÖTV", "base": None, "rate": None, "amount": round(data.sct_amount, 2), "formula": "doğrulanmış toplam tutar"})

    pre_vat_known = all(
        value is not None
        for value in (duty, additional, emy, data.anti_dumping_amount, kkdf, data.sct_amount, data.surveillance_unit_value)
    )
    vat_base = None
    vat = None
    if pre_vat_known:
        vat_base = customs_value + duty + additional + emy + data.anti_dumping_amount + kkdf + data.sct_amount + data.other_costs
        vat = percentage("vat", "KDV oranı", vat_base, data.vat_rate)
    else:
        if data.vat_rate is None:
            missing.append("KDV oranı")
        lines.append({"code": "vat", "label": "KDV", "base": None, "rate": data.vat_rate, "amount": None, "formula": "önceki vergi oranları eksik"})
    total = vat_base + vat if vat_base is not None and vat is not None else None
    unit = total / data.quantity if total is not None and data.quantity else None
    status: Literal["complete", "partial", "blocked"] = "complete" if total is not None and not missing else "partial"
    return LandedCostResult(
        status=status, currency=data.currency, lines=lines, customs_value=round(customs_value, 2),
        vat_base=round(vat_base, 2) if vat_base is not None else None,
        landed_total=round(total, 2) if total is not None else None,
        unit_landed_cost=round(unit, 4) if unit is not None else None,
        missing_rates=list(dict.fromkeys(missing)), warnings=warnings,
    )
